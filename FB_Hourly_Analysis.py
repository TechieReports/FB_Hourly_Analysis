import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Title and Description
st.title("Revenue and Spend Analysis Tool")
st.write("Upload multiple revenue files and a spend file to generate detailed reports.")

# File Uploads
revenue_files = st.file_uploader("Upload Revenue Files (CSV)", type=['csv'], accept_multiple_files=True)
spend_file = st.file_uploader("Upload Spend File (XLSX)", type=['xlsx'])

if revenue_files and spend_file:
    # Load and Merge Revenue Files
    revenue_data_list = [pd.read_csv(file) for file in revenue_files]
    revenue_data = pd.concat(revenue_data_list, ignore_index=True)

    # Load Spend Data
    spend_data = pd.read_excel(spend_file)

    # Data Cleaning for Spend File
    spend_data['campid'] = spend_data['Ad set name'].str.extract(r'\((\d+)\)').astype(float)
    spend_data['hour'] = spend_data['Time of day (ad account time zone)'].str.extract(r'(\d+):').astype(int)
    spend_cleaned = spend_data[['campid', 'hour', 'Amount spent (USD)', 'Results']].rename(
        columns={'Amount spent (USD)': 'hourly_spend', 'Results': 'hourly_results'}
    ).dropna()
    spend_cleaned['campid'] = spend_cleaned['campid'].astype(int)

    # Filter Revenue Data to Match Spend Campaigns
    matching_campaigns = spend_cleaned['campid'].unique()
    revenue_filtered = revenue_data[revenue_data['campid'].isin(matching_campaigns)]

    # Aggregate Revenue Data by Campaign and Hour
    revenue_filtered = revenue_filtered.groupby(['campid', 'hour'], as_index=False).agg({
        'estimated_earnings': 'sum',
        'clicks': 'sum'  # This represents total clicks for the hour
    }).rename(columns={'clicks': 'total_clicks'})

    # Aggregate Spend Data by Campaign and Hour
    spend_cleaned = spend_cleaned.groupby(['campid', 'hour'], as_index=False).agg({
        'hourly_spend': 'sum',
        'hourly_results': 'sum'
    })

    # Merge Aggregated Data
    merged_data = pd.merge(
        revenue_filtered,
        spend_cleaned,
        on=['campid', 'hour'],
        how='inner'
    )

    # Add Campaign Names
    campaign_name_mapping = spend_data[['campid', 'Ad set name']].drop_duplicates()
    campaign_name_mapping = campaign_name_mapping.rename(columns={'Ad set name': 'campaign_name'})
    merged_data = pd.merge(merged_data, campaign_name_mapping, on='campid', how='left')

    # Calculate Hourly Metrics
    merged_data['hourly_revenue'] = merged_data.groupby('campid')['estimated_earnings'].diff().fillna(merged_data['estimated_earnings'])
    merged_data['hourly_clicks'] = merged_data.groupby('campid')['total_clicks'].diff().fillna(merged_data['total_clicks'])

    # Calculate Hourly RPC (Revenue Per Click)
    merged_data['hourly_rpc'] = merged_data.apply(
        lambda row: row['hourly_revenue'] / row['hourly_clicks'] if row['hourly_clicks'] > 0 else 0, axis=1
    )

    # Calculate Hourly CPR (Cost Per Result)
    merged_data['hourly_cpr'] = merged_data.apply(
        lambda row: row['hourly_spend'] / row['hourly_results'] if row['hourly_results'] > 0 else 0, axis=1
    )

    # Calculate Profit, ROI, and Profit/Loss
    merged_data['profit'] = merged_data['hourly_revenue'] - merged_data['hourly_spend']
    merged_data['roi'] = merged_data.apply(
        lambda row: (row['profit'] / row['hourly_spend'] * 100) if row['hourly_spend'] > 0 else 0, axis=1
    )
    merged_data['profit_loss'] = merged_data['profit'].apply(
        lambda x: 'Profit' if x > 0 else 'Loss' if x < 0 else 'Break-Even'
    )

    # Round Decimals to 2 Places
    numeric_cols = ['hourly_spend', 'hourly_results', 'estimated_earnings', 'total_clicks',
                    'hourly_revenue', 'hourly_clicks', 'hourly_rpc', 'hourly_cpr', 'profit', 'roi']
    merged_data[numeric_cols] = merged_data[numeric_cols].round(2)

    # Reorder Columns
    merged_data = merged_data[[
        'campid', 'campaign_name', 'hour', 'hourly_spend', 'hourly_results',
        'estimated_earnings', 'total_clicks', 'hourly_revenue', 'hourly_clicks',
        'hourly_cpr', 'hourly_rpc', 'profit', 'roi', 'profit_loss'
    ]]

    # Generate Output File
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Metrics"

    # Write Headers
    headers = [col.upper() for col in merged_data.columns]
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write Data
    for row in dataframe_to_rows(merged_data, index=False, header=False):
        ws.append(row)

    # Color-Code Profit/Loss
    profit_loss_col = headers.index("PROFIT_LOSS") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=profit_loss_col, max_col=profit_loss_col):
        for cell in row:
            if cell.value == "Profit":
                cell.fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")  # Light green
            elif cell.value == "Loss":
                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")  # Light red
            elif cell.value == "Break-Even":
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow

    # Save Excel File to Buffer
    wb.save(output)
    output.seek(0)

    # Provide Download Link
    st.download_button(
        label="Download Report",
        data=output,
        file_name="Revenue_Spend_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
